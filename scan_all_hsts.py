import os
import sys
import struct
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse

class MasterHeader:
    def __init__(self):
        self.Title = None
        self.Version = None
        self.Max_nr_files = None
        self.Files_created = None
        self.Data_headers = []

class Header:
    def __init__(self):
        self.LogName = None
        self.SamplePeriod = None
        self.StartTime = None
        self.EndTime = None
        self.DataLength = None
        self.FilePointer = None
        self.Version = None

def clean_string(s):
    """Remove illegal Excel characters from string"""
    if not s:
        return ''
    # Remove control characters except tab, newline, carriage return
    # Then replace newlines/tabs with spaces
    import re
    # Remove control characters (0x00-0x1F except 0x09, 0x0A, 0x0D)
    s = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', s)
    # Replace remaining whitespace (tabs, newlines, etc.) with single space
    s = ' '.join(s.split())
    return s.strip()

def read_master_header(f):
    """Read the master header from HST file"""
    m = MasterHeader()
    raw_title = f.read(128).decode("cp1252", errors='ignore').rstrip("\x00")
    m.Title = clean_string(raw_title)
    f.read(8)  # ID
    f.read(2)  # Type
    m.Version = int.from_bytes(f.read(2), "little")
    f.read(4)  # Align
    f.read(4)  # Mode
    m.Max_nr_files = int.from_bytes(f.read(2), "little")
    m.Files_created = int.from_bytes(f.read(2), "little")
    f.read(24)  # Next, Addon, Align
    return m

def read_old_data_headers(m, f):
    """Read Version 5 (old type) data headers"""
    for _ in range(m.Files_created):
        f.read(144)  # Filename
        h = Header()
        h.Version = 5
        f.read(8)  # ID
        f.read(2)  # Type
        f.read(2)  # Version
        f.read(4)  # StartEvNo
        raw_logname = f.read(80).decode("cp1252", errors='ignore').rstrip("\x00")
        h.LogName = clean_string(raw_logname)
        f.read(4)  # Mode
        f.read(2)  # Area
        f.read(2)  # Priv
        f.read(2)  # FileType
        h.SamplePeriod = int.from_bytes(f.read(4), "little")
        f.read(8)  # EngUnits
        f.read(4)  # Format
        h.StartTime = datetime.fromtimestamp(int.from_bytes(f.read(4), "little"))
        h.EndTime = datetime.fromtimestamp(int.from_bytes(f.read(4), "little"))
        h.DataLength = int.from_bytes(f.read(4), "little")
        h.FilePointer = int.from_bytes(f.read(4), "little")
        f.read(6)  # EndEvNo, Align
        m.Data_headers.append(h)

def read_new_data_headers(m, f):
    """Read Version 6 (new type) data headers"""
    for _ in range(m.Files_created):
        f.read(272)  # Filename
        h = Header()
        h.Version = 6
        f.read(8)  # ID
        f.read(2)  # Type
        f.read(2)  # Version
        f.read(8)  # StartEvNo
        f.read(12)  # Align
        raw_logname = f.read(80).decode("cp1252", errors='ignore').rstrip("\x00")
        h.LogName = clean_string(raw_logname)
        f.read(4)  # Mode
        f.read(2)  # Area
        f.read(2)  # Priv
        f.read(2)  # FileType
        h.SamplePeriod = int.from_bytes(f.read(4), "little")
        f.read(8)  # EngUnits
        f.read(4)  # Format
        h.StartTime = datetime(1601, 1, 1) + timedelta(microseconds=int.from_bytes(f.read(8), "little") / 10)
        h.EndTime = datetime(1601, 1, 1) + timedelta(microseconds=int.from_bytes(f.read(8), "little") / 10)
        h.DataLength = int.from_bytes(f.read(4), "little")
        h.FilePointer = int.from_bytes(f.read(4), "little")
        f.read(14)  # EndEvNo, Align
        m.Data_headers.append(h)

def scan_hst_file(hst_path):
    """Extract essential metadata from HST file - optimized for speed"""
    try:
        # Get HST file stats (fast)
        file_stat = os.stat(hst_path)
        hst_size = file_stat.st_size
        hst_modified = datetime.fromtimestamp(file_stat.st_mtime)
        
        # Read headers only (fast - no data file reading)
        with open(hst_path, 'rb') as f:
            m = read_master_header(f)
            
            if m.Version == 5:
                read_old_data_headers(m, f)
            elif m.Version == 6:
                read_new_data_headers(m, f)
            else:
                return None
        
        # Get base info
        base_path = os.path.dirname(hst_path)
        base_name = os.path.splitext(os.path.basename(hst_path))[0]
        
        # Create one result row per data file
        results = []
        
        for idx, header in enumerate(m.Data_headers):
            # Find corresponding data file
            ext = f".{idx:03d}"
            data_file_path = os.path.join(base_path, base_name + ext)
            
            # Quick check if file exists
            data_exists = os.path.exists(data_file_path)
            if data_exists:
                data_size = os.path.getsize(data_file_path)
            else:
                data_size = 0
            
            # Calculate time span
            time_span = header.EndTime - header.StartTime
            time_span_days = time_span.total_seconds() / 86400
            
            # Calculate utilization
            utilization = (header.FilePointer / header.DataLength * 100) if header.DataLength > 0 else 0
            
            result = {
                # Essential HST info
                'directory': os.path.dirname(hst_path),
                'hst_filename': os.path.basename(hst_path),
                'sensor_name': m.Title,
                'hst_size_mb': hst_size / (1024 * 1024),
                'hst_modified': hst_modified,
                'version': m.Version,
                'sample_type': '2-byte' if m.Version == 5 else '8-byte',
                
                # Essential data file info
                'data_file_index': idx,
                'data_filename': os.path.basename(data_file_path),
                'data_exists': data_exists,
                'data_size_mb': data_size / (1024 * 1024),
                
                # Critical time info
                'start_time': header.StartTime,
                'end_time': header.EndTime,
                'time_span_days': time_span_days,
                
                # Essential sample statistics
                'valid_samples': header.FilePointer,
                'allocated_samples': header.DataLength,
                'utilization_pct': utilization,
                
                # Flags
                'has_issue': not data_exists or utilization < 50
            }
            
            results.append(result)
        
        return results
        
    except Exception as e:
        return [{
            'directory': os.path.dirname(hst_path),
            'hst_filename': os.path.basename(hst_path),
            'error': str(e)
        }]

def scan_directories(root_paths, progress_callback=None):
    """Scan all directories for HST files"""
    all_hst_files = []
    
    for root_path in root_paths:
        print(f"Scanning directory: {root_path}")
        for root, dirs, files in os.walk(root_path):
            for file in files:
                if file.upper().endswith('.HST'):
                    hst_path = os.path.join(root, file)
                    all_hst_files.append(hst_path)
    
    print(f"Found {len(all_hst_files)} HST files")
    
    results = []
    for i, hst_path in enumerate(all_hst_files, 1):
        if progress_callback:
            progress_callback(i, len(all_hst_files), hst_path)
        
        print(f"Processing [{i}/{len(all_hst_files)}]: {hst_path}")
        hst_results = scan_hst_file(hst_path)
        if hst_results:
            # hst_results is now a list of results (one per data file)
            results.extend(hst_results)
    
    return results

def create_excel_report(results, output_path):
    """Create simplified Excel report with essential columns only"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Files Detail"
    
    # Essential columns only
    headers = [
        'Directory',
        'HST Filename',
        'Sensor Name',
        'Data File Index',
        'Data Filename',
        'Data Exists',
        'Version',
        'Sample Type',
        'Start Time',
        'End Time',
        'Time Span (Days)',
        'Valid Samples',
        'Allocated Samples',
        'Utilization %',
        'HST Size (MB)',
        'Data Size (MB)',
        'HST Modified',
        'Has Issue',
        'Error'
    ]
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1).value = result.get('directory', '')
        ws.cell(row=row_idx, column=2).value = result.get('hst_filename', '')
        ws.cell(row=row_idx, column=3).value = result.get('sensor_name', '')
        ws.cell(row=row_idx, column=4).value = result.get('data_file_index', '')
        ws.cell(row=row_idx, column=5).value = result.get('data_filename', '')
        ws.cell(row=row_idx, column=6).value = 'Yes' if result.get('data_exists') else 'No'
        ws.cell(row=row_idx, column=7).value = result.get('version', '')
        ws.cell(row=row_idx, column=8).value = result.get('sample_type', '')
        ws.cell(row=row_idx, column=9).value = result.get('start_time')
        ws.cell(row=row_idx, column=10).value = result.get('end_time')
        ws.cell(row=row_idx, column=11).value = result.get('time_span_days', 0)
        ws.cell(row=row_idx, column=12).value = result.get('valid_samples', 0)
        ws.cell(row=row_idx, column=13).value = result.get('allocated_samples', 0)
        ws.cell(row=row_idx, column=14).value = result.get('utilization_pct', 0)
        ws.cell(row=row_idx, column=15).value = result.get('hst_size_mb', 0)
        ws.cell(row=row_idx, column=16).value = result.get('data_size_mb', 0)
        ws.cell(row=row_idx, column=17).value = result.get('hst_modified')
        ws.cell(row=row_idx, column=18).value = 'Yes' if result.get('has_issue') else 'No'
        ws.cell(row=row_idx, column=19).value = result.get('error', '')
        
        # Highlight rows with issues
        if result.get('has_issue'):
            for col in range(1, 20):
                ws.cell(row=row_idx, column=col).fill = PatternFill(start_color='FFF4CC', end_color='FFF4CC', fill_type='solid')
    
    # Format number columns
    for row in range(2, len(results) + 2):
        ws.cell(row=row, column=11).number_format = '0.00'  # Time Span
        ws.cell(row=row, column=12).number_format = '#,##0' # Valid Samples
        ws.cell(row=row, column=13).number_format = '#,##0' # Allocated Samples
        ws.cell(row=row, column=14).number_format = '0.00'  # Utilization
        ws.cell(row=row, column=15).number_format = '0.00'  # HST Size
        ws.cell(row=row, column=16).number_format = '0.00'  # Data Size
    
    # Format date columns
    for row in range(2, len(results) + 2):
        ws.cell(row=row, column=9).number_format = 'YYYY-MM-DD HH:MM:SS'   # Start Time
        ws.cell(row=row, column=10).number_format = 'YYYY-MM-DD HH:MM:SS'  # End Time
        ws.cell(row=row, column=17).number_format = 'YYYY-MM-DD HH:MM:SS'  # HST Modified
    
    # Auto-adjust column widths
    column_widths = {
        1: 60,  # Directory
        2: 35,  # HST Filename
        3: 50,  # Sensor Name
        4: 10,  # Data File Index
        5: 35,  # Data Filename
        6: 10,  # Data Exists
        7: 8,   # Version
        8: 12,  # Sample Type
        9: 20,  # Start Time
        10: 20, # End Time
        11: 15, # Time Span
        12: 15, # Valid Samples
        13: 17, # Allocated Samples
        14: 13, # Utilization
        15: 12, # HST Size
        16: 12, # Data Size
        17: 20, # HST Modified
        18: 10, # Has Issue
        19: 30  # Error
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze top row and first 3 columns
    ws.freeze_panes = 'D2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws['A1'] = 'HST Files Scan Summary'
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    # Count unique HST files
    unique_hst_files = len(set(r.get('hst_filename', '') for r in results if r.get('hst_filename')))
    total_data_files = len(results)
    files_with_errors = sum(1 for r in results if r.get('error'))
    files_with_issues = sum(1 for r in results if r.get('has_issue'))
    missing_data_files = sum(1 for r in results if not r.get('data_exists'))
    
    summary_data = [
        ['Total HST Files', unique_hst_files],
        ['Total Data Files (.00x)', total_data_files],
        ['Data Files with Issues', files_with_issues],
        ['Missing Data Files', missing_data_files],
        ['Files with Errors', files_with_errors],
        ['Version 5 Files', sum(1 for r in results if r.get('version') == 5)],
        ['Version 6 Files', sum(1 for r in results if r.get('version') == 6)],
        ['Total Valid Samples', sum(r.get('valid_samples', 0) for r in results)],
        ['Average Utilization %', sum(r.get('utilization_pct', 0) for r in results) / len(results) if results else 0]
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 3):
        summary_ws.cell(row=row_idx, column=1).value = label
        summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)
        summary_ws.cell(row=row_idx, column=2).value = value
        if isinstance(value, float):
            summary_ws.cell(row=row_idx, column=2).number_format = '0.00'
    
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 20
    
    wb.save(output_path)
    print(f"Report saved to: {output_path}")

def main():
    parser = argparse.ArgumentParser(description='Scan HST files and generate Excel report')
    parser.add_argument('directories', nargs='+', help='Root directories to scan for HST files')
    parser.add_argument('-o', '--output', default='hst_scan_report.xlsx', help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Validate directories
    for directory in args.directories:
        if not os.path.exists(directory):
            print(f"Error: Directory does not exist: {directory}")
            sys.exit(1)
    
    # Scan directories
    results = scan_directories(args.directories)
    
    if not results:
        print("No HST files found!")
        sys.exit(1)
    
    # Create Excel report
    create_excel_report(results, args.output)
    
    print(f"\nScan complete!")
    print(f"Total files scanned: {len(results)}")
    print(f"Report saved to: {args.output}")

if __name__ == "__main__":
    main()